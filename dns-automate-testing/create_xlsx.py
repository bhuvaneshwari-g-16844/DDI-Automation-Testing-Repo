from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_sheet(wb, sheet_name, headers, data):
    ws = wb.create_sheet(title=sheet_name)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    col_widths = [14, 8, 16, 18, 70, 70, 10, 16, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    return ws


headers = ["Date", "TC No", "Module", "Platform", "Test Case Description", "Expected Result", "Type", "Test Nature", "Status"]

dns_data = [
    ["18-03-2026", 1, "DNS Anomaly", "Windows & Linux", "Verify the total DNS anomaly count on the dashboard.", "Anomaly count should be accurate and reflect actual detected anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 2, "DNS Anomaly", "Windows & Linux", "Drill down each anomaly record and verify individual anomaly details.", "Each record's drill down must display correct anomaly type, domain, and timestamp.", "Positive", "Functional", ""],
    ["18-03-2026", 3, "DNS Anomaly", "Windows & Linux", "Check DNS anomaly data based on date filters: monthly, hourly, and daily.", "Correct anomaly data should load as per the selected filter.", "Positive", "Functional", ""],
    ["18-03-2026", 4, "DNS Anomaly", "Windows & Linux", 'Check anomaly data for "Today", "Yesterday", and "Custom Date".', "Data must match the selected date range accurately.", "Positive", "Functional", ""],
    ["18-03-2026", 5, "DNS Anomaly", "Windows & Linux", "Verify anomaly detection for DNS tunneling attempts.", "DNS tunneling anomalies should be detected and displayed with correct details.", "Positive", "Functional", ""],
    ["18-03-2026", 6, "DNS Anomaly", "Windows & Linux", "Verify anomaly detection for DNS amplification/DDoS patterns.", "Amplification attack patterns should be flagged and recorded as anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 7, "DNS Anomaly", "Windows & Linux", "Verify anomaly detection for unusually high query volume from a single source IP.", "High query volume anomalies should be detected and flagged with source IP details.", "Positive", "Functional", ""],
    ["18-03-2026", 8, "DNS Anomaly", "Windows & Linux", "Verify anomaly detection for NXDOMAIN flood patterns.", "Excessive NXDOMAIN responses should be flagged as anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 9, "DNS Anomaly", "Windows & Linux", "Verify anomaly detection for queries to known malicious/suspicious domains.", "Queries to suspicious domains should be flagged with appropriate severity level.", "Positive", "Functional", ""],
    ["18-03-2026", 10, "DNS Anomaly", "Windows & Linux", "Check zone-level anomaly filtering based on scope.", "Zone anomalies should display correctly when filtered by scope.", "Positive", "Functional", ""],
    ["18-03-2026", 11, "DNS Anomaly", "Windows & Linux", "Verify anomaly severity classification (Critical, High, Medium, Low).", "Each anomaly should be categorized with the correct severity level.", "Positive", "Functional", ""],
    ["18-03-2026", 12, "DNS Anomaly", "Windows & Linux", "Verify anomaly data for hosted, non-hosted, and blocked domains with scope filtering.", "Scope filters should apply correctly and display accurate anomaly data for each category.", "Positive", "Functional", ""],
    ["18-03-2026", 13, "DNS Anomaly", "Windows & Linux", "Verify that anomaly alerts/notifications are triggered for critical anomalies.", "Alerts should be generated and displayed for critical severity anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 14, "DNS Anomaly", "Windows & Linux", "Verify export of DNS anomaly data in PDF and CSV formats.", "Both PDF and CSV files should download correctly and contain complete anomaly data.", "Positive", "Functional", ""],
    ["18-03-2026", 15, "DNS Anomaly", "Windows & Linux", "Verify anomaly detection for record type-based filtering (A, AAAA, MX, PTR, etc.).", "Results must reflect only anomalies for the selected record type with correct counts.", "Positive", "Functional", ""],
    ["18-03-2026", 16, "DNS Anomaly", "Windows & Linux", "Check RCODE-based anomaly detection (SERVFAIL, REFUSED, FORMERR, etc.).", "Anomalies should be categorized correctly by response code.", "Positive", "Functional", ""],
    ["18-03-2026", 17, "DNS Anomaly", "Windows & Linux", "Verify index page search by domain name and drill down to detailed anomaly pages.", "All drill down views should load correct domain-specific anomaly data.", "Positive", "Functional", ""],
    ["18-03-2026", 18, "DNS Anomaly", "Windows & Linux", "Verify index page search by IP address and validate all anomaly drill down views.", "All associated anomalies must load and match the searched IP context.", "Positive", "Functional", ""],
    ["18-03-2026", 19, "DNS Anomaly", "Windows & Linux", "Verify anomaly trend graph/chart accuracy over selected time range.", "Graph should reflect correct anomaly trends matching the table data.", "Positive", "Functional", ""],
    ["18-03-2026", 20, "DNS Anomaly", "Windows & Linux", "Validate anomaly correlation between multiple related DNS events.", "Related anomalies should be grouped or linked for easier investigation.", "Positive", "Functional", ""],
    ["18-03-2026", 21, "DNS Anomaly", "Windows & Linux", "Drill down all pages in DNS anomaly records and validate anomaly counts.", "All pages should load with correct anomaly data and counts.", "Positive", "Regression", ""],
    ["18-03-2026", 22, "DNS Anomaly", "Windows & Linux", "Compare exported anomaly data (PDF/CSV) with UI table for accuracy.", "Exported files must match on-screen anomaly data exactly.", "Positive", "Regression", ""],
    ["18-03-2026", 23, "DNS Anomaly", "Windows & Linux", "Test loading of additional anomaly records via pagination or infinite scroll.", "All records must load without break or duplication.", "Positive", "Regression", ""],
    ["18-03-2026", 24, "DNS Anomaly", "Windows & Linux", 'Validate functionality of "Reset Filters" button on anomaly page.', "All filters should clear and default anomaly view should reload.", "Positive", "Regression", ""],
    ["18-03-2026", 25, "DNS Anomaly", "Linux", "Verify anomaly count views based on zone after re-testing previous fixes.", "Anomaly count should match zone-specific data.", "Positive", "Regression", ""],
    ["18-03-2026", 26, "DNS Anomaly", "Windows & Linux", "Validate system behavior with bulk anomaly entries and cross-check in PostgreSQL.", "System should handle bulk anomaly data smoothly; database and UI counts must match.", "Positive", "Performance", ""],
    ["18-03-2026", 27, "DNS Anomaly", "Windows & Linux", "Monitor Task Manager during bulk anomaly detection for performance metrics (CPU/RAM).", "System resources should remain within acceptable limits without major slowdown.", "Positive", "Performance", ""],
    ["18-03-2026", 28, "DNS Anomaly", "Windows & Linux", "Check if anomaly data updates in near real-time (if live tracking is enabled).", "New anomalies should appear without full page reload if feature is active.", "Positive", "Performance", ""],
    ["18-03-2026", 29, "DNS Anomaly", "Windows & Linux", "Verify anomaly timestamps respect system time zone settings.", "Anomaly data must reflect accurate local time zones.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 30, "DNS Anomaly", "Windows & Linux", "Enter invalid or out-of-bound date ranges and check system response.", "System should show validation messages and block invalid requests.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 31, "DNS Anomaly", "Windows & Linux", "Apply filters that yield no anomaly data and observe UI behavior.", 'UI should display appropriate "No Anomaly Data Found" message.', "Negative", "Non-Functional", ""],
    ["18-03-2026", 32, "DNS Anomaly", "Windows & Linux", "Simulate delay or loss of sync between agent and backend for anomaly data.", "System should retry, log issue, or show warning if anomaly data is not syncing.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 33, "DNS Anomaly", "Windows & Linux", "Verify DNS anomaly access for admin vs. read-only roles.", "Only authorized users should access or export anomaly data.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 34, "DNS Anomaly", "Windows & Linux", "Check if old anomaly data beyond retention period is purged automatically.", "Data older than policy-defined limits should be removed securely.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 35, "DNS Anomaly", "Windows & Linux", "Test anomaly UI in different screen sizes or window resizing.", "UI should adjust layout without data loss or misalignment.", "Negative", "Non-Functional", ""],
]

dhcp_data = [
    ["18-03-2026", 1, "DHCP Anomaly", "Windows & Linux", "Verify the total DHCP anomaly count on the dashboard.", "Anomaly count should be accurate and reflect actual detected DHCP anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 2, "DHCP Anomaly", "Windows & Linux", "Drill down each DHCP anomaly record and verify individual anomaly details.", "Each record's drill down must display correct anomaly type, IP, MAC address, and timestamp.", "Positive", "Functional", ""],
    ["18-03-2026", 3, "DHCP Anomaly", "Windows & Linux", "Check DHCP anomaly data based on date filters: monthly, hourly, and daily.", "Correct anomaly data should load as per the selected filter.", "Positive", "Functional", ""],
    ["18-03-2026", 4, "DHCP Anomaly", "Windows & Linux", 'Check DHCP anomaly data for "Today", "Yesterday", and "Custom Date".', "Data must match the selected date range accurately.", "Positive", "Functional", ""],
    ["18-03-2026", 5, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for DHCP starvation/exhaustion attacks.", "DHCP starvation anomalies should be detected and flagged with source details.", "Positive", "Functional", ""],
    ["18-03-2026", 6, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for rogue DHCP server presence.", "Rogue DHCP server anomalies should be detected and alerted immediately.", "Positive", "Functional", ""],
    ["18-03-2026", 7, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for unusually high DHCP request volume from a single MAC address.", "High request volume anomalies should be flagged with MAC address details.", "Positive", "Functional", ""],
    ["18-03-2026", 8, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for DHCP lease conflicts (duplicate IP assignment).", "IP conflict anomalies should be detected and displayed with both conflicting records.", "Positive", "Functional", ""],
    ["18-03-2026", 9, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for unauthorized/unknown MAC addresses requesting leases.", "Unknown MAC address lease requests should be flagged as anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 10, "DHCP Anomaly", "Windows & Linux", "Check scope-level DHCP anomaly filtering.", "Scope-based anomalies should display correctly when filtered.", "Positive", "Functional", ""],
    ["18-03-2026", 11, "DHCP Anomaly", "Windows & Linux", "Verify anomaly severity classification (Critical, High, Medium, Low) for DHCP events.", "Each DHCP anomaly should be categorized with the correct severity level.", "Positive", "Functional", ""],
    ["18-03-2026", 12, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for rapid lease acquire/release cycles (DHCP flooding).", "Rapid DHCP flooding patterns should be detected and flagged.", "Positive", "Functional", ""],
    ["18-03-2026", 13, "DHCP Anomaly", "Windows & Linux", "Verify that anomaly alerts/notifications are triggered for critical DHCP anomalies.", "Alerts should be generated and displayed for critical DHCP severity anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 14, "DHCP Anomaly", "Windows & Linux", "Verify export of DHCP anomaly data in PDF and CSV formats.", "Both PDF and CSV files should download correctly with complete DHCP anomaly data.", "Positive", "Functional", ""],
    ["18-03-2026", 15, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for expired leases that are still being used.", "Expired lease usage anomalies should be flagged with relevant IP and MAC details.", "Positive", "Functional", ""],
    ["18-03-2026", 16, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for DHCP DECLINE and DHCP NAK message patterns.", "DECLINE/NAK patterns should be flagged as potential anomalies.", "Positive", "Functional", ""],
    ["18-03-2026", 17, "DHCP Anomaly", "Windows & Linux", "Verify index page search by IP address and drill down to DHCP anomaly details.", "All drill down views should load correct IP-specific DHCP anomaly data.", "Positive", "Functional", ""],
    ["18-03-2026", 18, "DHCP Anomaly", "Windows & Linux", "Verify index page search by MAC address and validate drill down views.", "All associated DHCP anomalies must load and match the searched MAC context.", "Positive", "Functional", ""],
    ["18-03-2026", 19, "DHCP Anomaly", "Windows & Linux", "Verify DHCP anomaly trend graph/chart accuracy over selected time range.", "Graph should reflect correct anomaly trends matching the table data.", "Positive", "Functional", ""],
    ["18-03-2026", 20, "DHCP Anomaly", "Windows & Linux", "Verify anomaly detection for subnet/pool exhaustion thresholds.", "Anomaly should be raised when pool utilization exceeds defined threshold.", "Positive", "Functional", ""],
    ["18-03-2026", 21, "DHCP Anomaly", "Windows & Linux", "Cross-check DHCP anomaly data with DHCP lease records table.", "Anomaly records should correlate with corresponding lease entries.", "Positive", "Regression", ""],
    ["18-03-2026", 22, "DHCP Anomaly", "Windows & Linux", "Drill down all pages in DHCP anomaly records and validate counts.", "All pages should load with correct anomaly data and counts.", "Positive", "Regression", ""],
    ["18-03-2026", 23, "DHCP Anomaly", "Windows & Linux", "Compare exported DHCP anomaly data (PDF/CSV) with UI table for accuracy.", "Exported files must match on-screen DHCP anomaly data exactly.", "Positive", "Regression", ""],
    ["18-03-2026", 24, "DHCP Anomaly", "Windows & Linux", "Test loading of additional DHCP anomaly records via pagination or infinite scroll.", "All records must load without break or duplication.", "Positive", "Regression", ""],
    ["18-03-2026", 25, "DHCP Anomaly", "Windows & Linux", 'Validate functionality of "Reset Filters" button on DHCP anomaly page.', "All filters should clear and default view should reload.", "Positive", "Regression", ""],
    ["18-03-2026", 26, "DHCP Anomaly", "Windows & Linux", "Validate system behavior with bulk DHCP anomaly entries and cross-check in PostgreSQL.", "System should handle bulk data smoothly; database and UI counts must match.", "Positive", "Performance", ""],
    ["18-03-2026", 27, "DHCP Anomaly", "Windows & Linux", "Monitor Task Manager during bulk DHCP anomaly detection for performance metrics.", "System resources should remain within acceptable limits without major slowdown.", "Positive", "Performance", ""],
    ["18-03-2026", 28, "DHCP Anomaly", "Windows & Linux", "Check if DHCP anomaly data updates in near real-time.", "New DHCP anomalies should appear without full page reload if feature is active.", "Positive", "Performance", ""],
    ["18-03-2026", 29, "DHCP Anomaly", "Windows & Linux", "Verify DHCP anomaly timestamps respect system time zone settings.", "Anomaly data must reflect accurate local time zones.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 30, "DHCP Anomaly", "Windows & Linux", "Enter invalid or out-of-bound date ranges and check system response.", "System should show validation messages and block invalid requests.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 31, "DHCP Anomaly", "Windows & Linux", "Apply filters that yield no DHCP anomaly data and observe UI behavior.", 'UI should display appropriate "No DHCP Anomaly Data Found" message.', "Negative", "Non-Functional", ""],
    ["18-03-2026", 32, "DHCP Anomaly", "Windows & Linux", "Simulate delay or loss of sync between DHCP agent and backend for anomaly data.", "System should retry, log issue, or show warning if data is not syncing.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 33, "DHCP Anomaly", "Windows & Linux", "Verify DHCP anomaly access for admin vs. read-only roles.", "Only authorized users should access or export DHCP anomaly data.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 34, "DHCP Anomaly", "Windows & Linux", "Check if old DHCP anomaly data beyond retention period is purged automatically.", "Data older than policy-defined limits should be removed securely.", "Negative", "Non-Functional", ""],
    ["18-03-2026", 35, "DHCP Anomaly", "Windows & Linux", "Test DHCP anomaly UI in different screen sizes or window resizing.", "UI should adjust layout without data loss or misalignment.", "Negative", "Non-Functional", ""],
]

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

create_sheet(wb, "DNS Anomaly", headers, dns_data)
create_sheet(wb, "DHCP Anomaly", headers, dhcp_data)

output_path = "/home/bhuva-16844/Documents/playwright-python/dns-playwright-framework/anomaly_testcases.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
